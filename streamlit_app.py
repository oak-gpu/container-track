import io
import re
from datetime import datetime

import openpyxl
import streamlit as st

import config
from trackers import ais as tracker_ais


# ── Column auto-detection ─────────────────────────────────────────────────────

_HEADER_MAP = {
    "vessel":          ["VESSEL / VOYAGE NAME", "VESSEL NAME", "VESSEL"],
    "carrier":         ["CARRIER", "FORWARDER / LINE", "FORWARDER"],
    "etd":             ["ETD"],
    "eta":             ["ETA"],
    "container":       ["CONTAINER ID", "CONTAINER NO", "CONTAINER"],
    "location":        ["CONTAINER LOCATION"],
    "actual_delivery": ["ACTUAL DELIVERY DATE"],
    "mmsi":            ["MMSI"],
    "updated_eta":     ["UPDATED ETA"],
    "actual_arr":      ["ACTUAL ARRIVAL"],
}


def detect_columns(ws) -> dict:
    """Read header row and return {key: column_index} for all known columns."""
    cols = {}
    for cell in ws[1]:
        if not cell.value:
            continue
        val = str(cell.value).strip().upper()
        for key, options in _HEADER_MAP.items():
            if key not in cols and val in [o.upper() for o in options]:
                cols[key] = cell.column

    # output columns that don't exist yet go after the last used column
    next_col = ws.max_column + 1
    for key in ("mmsi", "updated_eta", "actual_arr"):
        if key not in cols:
            cols[key] = next_col
            next_col += 1

    return cols

# Load API key from Streamlit secrets (set via Streamlit Cloud dashboard)
try:
    config.MYSHIPTRACKING_API_KEY = st.secrets["MYSHIPTRACKING_API_KEY"]
except Exception:
    pass

# ── Port name lookup ──────────────────────────────────────────────────────────

_PORT_NAMES = {
    "USSAV": "Savannah, GA",     "USLAX": "Los Angeles, CA",
    "USLGB": "Long Beach, CA",   "USNYC": "New York, NY",
    "USHOU": "Houston, TX",      "USORF": "Norfolk, VA",
    "USCHA": "Charleston, SC",   "USJAX": "Jacksonville, FL",
    "USMOB": "Mobile, AL",       "USBAL": "Baltimore, MD",
    "USSEA": "Seattle, WA",      "USOAK": "Oakland, CA",
    "USTAC": "Tacoma, WA",       "USBOS": "Boston, MA",
    "USNOL": "New Orleans, LA",
    "MTMLA": "Valletta, Malta",  "MAPTM": "Tanger Med, Morocco",
    "ITSAL": "Salerno, Italy",   "ITGOA": "Genoa, Italy",
    "ITCAG": "Cagliari, Italy",  "ITTRS": "Trieste, Italy",
    "ESBCN": "Barcelona, Spain", "ESVLC": "Valencia, Spain",
    "ESLPA": "Las Palmas",       "GRSKG": "Thessaloniki, Greece",
    "GRPIR": "Piraeus, Greece",  "TRIZM": "Izmir, Turkey",
    "TRGEM": "Gemlik, Turkey",   "TRMER": "Mersin, Turkey",
    "TRAMS": "Ambarli, Turkey",  "TRIST": "Istanbul, Turkey",
    "AEJEA": "Jebel Ali, UAE",   "EGPSD": "Port Said, Egypt",
    "EGALS": "Alexandria, Egypt","SAJED": "Jeddah, Saudi Arabia",
    "SGSIN": "Singapore",        "CNSHA": "Shanghai, China",
    "CNNGB": "Ningbo, China",    "HKHKG": "Hong Kong",
    "PKKAR": "Karachi, Pakistan","INBOM": "Mumbai, India",
}


def _locode_to_name(code: str) -> str:
    if not code:
        return ""
    return _PORT_NAMES.get(code.strip().upper(), code.strip())


def _eta_date(eta: str) -> str:
    if not eta:
        return ""
    return eta.replace("T", " ").split(" ")[0].split(".")[0]


def _fmt_date(iso_date: str) -> str:
    if not iso_date:
        return ""
    try:
        dt = datetime.strptime(iso_date, "%Y-%m-%d")
        return f"{dt.day} {dt.strftime('%b %Y')}"
    except Exception:
        return iso_date


def _build_situation(ais: dict) -> tuple[str, str]:
    from trackers.ais import is_us_port, _port_name, _port_locode

    nav          = ais.get("nav_status", "")
    dest         = ais.get("destination", "")
    eta          = ais.get("eta", "")
    current_port = ais.get("current_port", "")
    next_port    = ais.get("next_port", "")
    next_eta     = ais.get("next_port_eta", "")
    arrived_at   = _eta_date(ais.get("arrived_at", ""))

    cur_locode  = _port_locode(current_port)
    cur_name    = _locode_to_name(_port_name(current_port) or cur_locode)
    next_locode = _port_locode(next_port)
    dest_name   = _locode_to_name(dest)

    if is_us_port(cur_locode) or (nav in ("Moored", "At anchor", "At Anchor") and is_us_port(dest)):
        port = cur_name if is_us_port(cur_locode) else dest_name
        actual = arrived_at or _eta_date(eta)
        return f"ARRIVED — {port}", actual

    if is_us_port(dest):
        eta_date = _eta_date(eta)
        status = f"IN TRANSIT → {dest_name}"
        if eta_date:
            status += f" | ETA {_fmt_date(eta_date)}"
        return status, eta_date

    if is_us_port(next_locode):
        eta_date = _eta_date(next_eta)
        next_name = _locode_to_name(_port_name(next_port) or next_locode)
        stop = cur_name or dest_name
        status = f"AT PORT: {stop} → Next: {next_name}"
        if eta_date:
            status += f" | US ETA {_fmt_date(eta_date)}"
        return status, eta_date

    if nav in ("Moored", "At anchor", "At Anchor"):
        port_info = cur_name or dest_name or "intermediate port"
        return f"AT PORT: {port_info} (en route to US)", ""

    if dest:
        eta_date = _eta_date(eta)
        status = f"AT SEA → {dest_name}"
        if eta_date:
            status += f" | Arrives {_fmt_date(eta_date)}"
        return status, ""

    return "AT SEA (destination unknown)", ""


# ── Page config ───────────────────────────────────────────────────────────────

st.set_page_config(page_title="Container Tracker", page_icon="🚢", layout="centered")
st.title("🚢 Container Tracker")
st.caption("Upload your Excel file to track shipments via AIS. Original columns are never modified.")

# ── File upload ───────────────────────────────────────────────────────────────

uploaded = st.file_uploader("Select your Excel file (.xlsx)", type=["xlsx"])

if not uploaded:
    st.stop()

if not st.button("Track Containers", type="primary", use_container_width=True):
    st.stop()

# ── Processing ────────────────────────────────────────────────────────────────

wb = openpyxl.load_workbook(uploaded)
total = skipped = 0

tracker_ais.clear_run_cache()

with st.status("Tracking containers…", expanded=True) as status_box:
    for ws in wb.worksheets:
        cols = detect_columns(ws)

        # skip sheets that don't look like tracking sheets
        missing = [k for k in ("vessel", "container", "location") if k not in cols]
        if missing:
            st.write(f"**Sheet '{ws.title}'** — skipped (missing columns: {', '.join(missing)})")
            continue

        st.write(f"### Sheet: {ws.title}")

        for key, label in [("mmsi", "MMSI"), ("updated_eta", "UPDATED ETA"), ("actual_arr", "ACTUAL ARRIVAL")]:
            if not ws.cell(row=1, column=cols[key]).value:
                ws.cell(row=1, column=cols[key]).value = label

        for row_idx in range(config.HEADER_ROWS + 1, ws.max_row + 1):
            carrier      = ws.cell(row=row_idx, column=cols.get("carrier", 1)).value
            container_id = ws.cell(row=row_idx, column=cols["container"]).value

            if not carrier and not container_id:
                break
            if not container_id:
                continue

            carrier      = str(carrier or "").strip()
            container_id = str(container_id).strip().upper()
            total += 1

            # skip if manually delivered
            delivered = ws.cell(row=row_idx, column=cols["actual_delivery"]).value if "actual_delivery" in cols else None
            if delivered:
                skipped += 1
                continue

            # skip if already marked arrived
            current_loc = str(ws.cell(row=row_idx, column=cols["location"]).value or "")
            if current_loc.startswith("ARRIVED"):
                skipped += 1
                continue

            st.write(f"**{container_id}** ({carrier})")

            vessel_raw  = ws.cell(row=row_idx, column=cols["vessel"]).value if "vessel" in cols else None
            mmsi_cell   = ws.cell(row=row_idx, column=cols["mmsi"])
            eta_cell    = ws.cell(row=row_idx, column=cols["updated_eta"])
            actual_cell = ws.cell(row=row_idx, column=cols["actual_arr"])
            loc_cell    = ws.cell(row=row_idx, column=cols["location"])
            etd_cell    = ws.cell(row=row_idx, column=cols["etd"]) if "etd" in cols else None
            booked_eta  = ws.cell(row=row_idx, column=cols["eta"]) if "eta" in cols else None

            if not vessel_raw:
                st.write("  ↳ no vessel name")
                continue

            mmsi, err = tracker_ais.resolve_mmsi(str(vessel_raw), mmsi_cell.value)
            if not mmsi:
                st.write(f"  ↳ AIS: {err}")
                continue

            mmsi_cell.value = mmsi
            ais_data = tracker_ais.get_vessel_status(mmsi)

            if ais_data.get("error"):
                err_msg = ais_data["error"]
                if "404" in err_msg:
                    st.write(f"  ↳ vessel not found in AIS coverage")
                else:
                    st.write(f"  ↳ AIS error: {err_msg}")
                continue

            situation, us_eta = _build_situation(ais_data)
            loc_cell.value  = situation
            eta_cell.value  = us_eta

            if situation.startswith("ARRIVED") and us_eta:
                actual_cell.value = us_eta
                etd  = _fmt_date(_eta_date(str(etd_cell.value if etd_cell else "")))
                exp  = _fmt_date(_eta_date(str(booked_eta.value if booked_eta else "")))
                act  = _fmt_date(us_eta)
                st.write(f"  ↳ {situation} | ETD {etd} | Expected {exp} | Actual {act}")
            else:
                st.write(f"  ↳ {situation}")

    status_box.update(
        label=f"Done — {total} containers processed, {skipped} skipped",
        state="complete",
    )

# ── Download ──────────────────────────────────────────────────────────────────

output = io.BytesIO()
wb.save(output)
output.seek(0)

timestamp = datetime.now().strftime("%Y-%m-%d %H-%M")
orig_name = uploaded.name.replace(".xlsx", "")
out_name  = f"{orig_name} - Processed {timestamp}.xlsx"

st.success(f"✓ {total} containers processed, {skipped} skipped")
st.download_button(
    label="⬇️ Download processed file",
    data=output,
    file_name=out_name,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)
